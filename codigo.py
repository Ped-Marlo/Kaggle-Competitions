'''
V2-takes the q values from an Excel file, (line 159) and classifies which files dont have the q specified
V3-attemps to create a model for each point of the net
V4-implementation of sections properly(no tolerance)
V5-ONE SINGLE "same" mesh/ eliminate the x y z dependency doing model by points

'''

#from sklearn.metrics import classification_report, confusion_matrix
import pandas as pd, numpy as np, matplotlib, matplotlib.pyplot as plt ,matplotlib.cm as cmx, time
#from sklearn.metrics import mean_squared_error as mse
from mpl_toolkits.mplot3d import Axes3D
import glob
#from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
#from sklearn.model_selection import GridSearchCV

#from sklearn.linear_model import LogisticRegression#categorical data
#from sklearn import preprocessing
#from sklearn import linear_model
#from sklearn.model_selection import cross_val_score
#from pandas  import scatter_matrix
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LassoCV
#from sklearn.linear_model import Lasso
from sklearn.linear_model import RidgeCV, BayesianRidge
from sklearn.linear_model import LassoLarsCV, LassoLarsIC, LarsCV
from sklearn.linear_model import ElasticNetCV
#from sklearn.pipeline import Pipeline
from sklearn import linear_model
from sklearn.linear_model import LinearRegression as LR
#from sklearn import svm
#from sklearn.metrics import  r2_score
from sklearn.linear_model.stochastic_gradient import SGDRegressor
from sklearn.svm import LinearSVR
#import keras
from sklearn.feature_selection import f_regression
import os.path
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
#import plotly.graph_objects as go
#from tkinter import  Tk, Label, Button, Scale, Canvas, Frame
from tkinter import *
from tkinter import ttk
import copy
from PIL import ImageTk, Image
import openpyxl
from openpyxl import load_workbook



#==============================================================================
START=time.time()

def scatter4d( x, y, z, cs, title,colorsMap='rainbow'): #jet tab20c for discrete range
    cm = plt.get_cmap(colorsMap)
#    cNorm = matplotlib.colors.Normalize(vmin=-1.4, vmax=0.4)
    cNorm = matplotlib.colors.Normalize(vmin=min(cs), vmax=max(cs))
    scalarMap = cmx.ScalarMappable(norm=cNorm, cmap=cm)
    fig = plt.figure()
    ax = Axes3D(fig)
    ax.scatter(x, y, z, c=scalarMap.to_rgba(cs))
    scalarMap.set_array(cs)
    fig.colorbar(scalarMap,label='Cp') ##DeltaP (daN/m2)
    ax.set_xlabel('x',fontsize=14, fontweight='bold')
    ax.set_ylabel('y',fontsize=14, fontweight='bold')
    ax.set_zlabel('z',fontsize=14, fontweight='bold')
    ax.set_title(title,fontsize=10,fontweight='bold')


    max_range = np.array([max(x)-min(x), max(y)-min(y), max(z)-min(z)]).max() / 2.0
    mid_x = (max(x)+min(x)) * 0.5
    mid_y = (max(y)+min(y)) * 0.5
    mid_z = (max(z)+min(z)) * 0.5
    ax.set_xlim(mid_x - max_range, mid_x + max_range)
    ax.set_ylim(mid_y - max_range, mid_y + max_range)
    ax.set_zlim(mid_z - max_range, mid_z + max_range)



def DataFrame_Creator():  #takes a file.txt( that includes the flight parameters Mach number & angle of attack AoA ) with geometry and pressure coefficients inputs (with different format depending of the way the mesh was originated ) and organizes it into a DataFrame structure. It looks for the flight parameters on an excel file and checks that the dynamic pressure q is available for all of them, then it attaches the q to the dataframe is it exist and delete the case if it doesn't

     def Txt_Reader(fname): #preparing and organizing the data from the text file
         f = open(fname,'r')
         Lines = f.readlines()
         f.close()
         Lines.append( '\n')
         zones = []
         zones.extend( i for i in Lines if i.startswith('ZONE '))
         zones = [ [ z.split('="')[1] for z  in zones ] [j].rstrip('"\n') for j,t in enumerate(zones) ]

         Nodes_elemts = []
         Nodes_elemts.extend(i for i  in Lines if i.startswith(' Nodes'))
         elements = {}
         nodes = {}
         Pos_zones_lines = [ k+1 for k,l in enumerate(Lines) if l.startswith(' DT') ]
         Datablock = [ k for k,l in enumerate(Lines) if l.startswith(' DATAPACKING') ]

         for T,U in enumerate(Nodes_elemts):
             elements[zones[T]] = int(Nodes_elemts [T].split('=')[2].split(',')[0])
             nodes[zones[T]]=int(Nodes_elemts [T].split('=')[1].split(',')[0])

         Array = {}
         if ' DATAPACKING=BLOCK\n' in Lines[Datablock[0]] and nodes:#uses the inherent property of a list, if isempty== there is not nodes
             for N,Zone in enumerate(zones):
                 Count = Pos_zones_lines[N]
                 Array[Zone] = {}
                 for Dim in ['x','y','z','cp']:#,'connectivity']:
                     Array[Zone][Dim] = []
                     while Count < len(Lines) and len(Array[Zone][Dim]) < nodes[Zone]:
         #                    if Dim is 'connectivity': #funciona solo con DT=single
         #                        for Elx in np.arange(elements[Zone]):
         #                            array[Zone][Dim].append(list(set(map(int,Lines[Count+Elx].split()))))
         #                        break
         #                    else:
                             Array[Zone][Dim].extend(list(map(float,Lines[Count].split())))
                             Count+=1

         elif ' DATAPACKING=BLOCK\n'in Lines[Datablock[0]]  and  not nodes:
             Data={}
             for Pos,Zon in enumerate(zones):
                 Count= Pos_zones_lines[Pos]
                 Data[Zon] = []

                 while (not Lines[Count].startswith('ZONE ') and (Count)<len(Lines)-1):
                     Data[Zon].extend(Lines[Count].split())
                     Count+= 1

                 Data[Zon] = [float(i) for i in Data[Zon]]

                 Array[Zon] = {}
                 Array[Zon]['x'] = Data[Zon][:int(len(Data[Zon])/4)]
                 Array[Zon]['y'] = Data[Zon][int(len(Data[Zon])/4):int(len(Data[Zon])*2/4)]
                 Array[Zon]['z'] = Data[Zon][int(len(Data[Zon])*2/4):int(len(Data[Zon])*3/4)]
                 Array[Zon]['cp']= Data[Zon][int(len(Data[Zon])*3/4):]

         elif ' DATAPACKING=POINT\n'in Lines[Datablock[0]] and  not nodes:
             for Pos,Zon in enumerate(zones):
                 Array[Zon] = {}

                 for ind,Dim in enumerate (['x','y','z','cp']):
                     Count = Pos_zones_lines[Pos]
                     Array[Zon][Dim] = []

                     while (not Lines[Count].startswith('ZONE ') and (Count)<len(Lines)-1):
                         Array[Zon][Dim].append(format(Lines[Count].split()[ind], '.10f'))
                         Count+=1

         elif ' DATAPACKING=POINT\n' in Lines[Datablock[0]] and   nodes:
              for N,Zone in enumerate(zones):
                  Array[Zone]={}

                  for ind, Dim in enumerate(['x','y','z','cp']):
                     Count= Pos_zones_lines[N]
                     Array[Zone][Dim] = []

                     while Count<len(Lines) and len(Array[Zone][Dim])<nodes[Zone]:
                         Array[Zone][Dim].append(float(Lines[Count].split()[ind]))
                         Count+=1

         CP=[]
         X=[]
         Y=[]
         Z=[]
         for block,dat in Array.items():
             for dim,val in dat.items():
                 if dim=='cp':
                     CP.extend(val)
                 elif dim=='x':
                     X.extend(val)
                 elif dim=='y':
                     Y.extend(val)
                 elif dim=='z':
                     Z.extend(val)

         return   CP,X,Y,Z,Array #elements,nodes,

     if os.path.exists(r"CSV350"):
          df = pd.read_csv(r"CSV350")#readimg the file like this 2 seconds( with thte first subroutine 10s)

     else:

          Txtfiles = []
          for File in glob.glob("*.dat"):
              Txtfiles.append(File)

          Txtfiles2 = copy.deepcopy(Txtfiles)#used for knowing which files are missing

          FILE = r'V010RP1804776_v1.0_A350-XWB_BF_v76_PiP_FTVL_Pressure_Mechanical_Cases_11052018_v4.xlsx'
          DF_excel_reader = pd.read_excel (FILE, sheet_name='Info',skiprows=[0,1,2])
          Q_excel_list = list(DF_excel_reader['Unnamed: 3'][55:126])
          Case = list(DF_excel_reader['Unnamed: 2'][55:126])

          CaseNumeric = []
          for F in Case:
                MachC = (float(F.split('_')[0].upper().lstrip('M')))/100

                if F.split('_')[1].lower().startswith('an'):
                    AlphaC = -(float(F.split('_')[1].lstrip('an')))/100
                elif F.split('_')[1].lower().startswith('a'):
                    AlphaC = (float(F.split('_')[1].lstrip('a')))/100

                if F.split('_')[2].lower().startswith('bn'):
                    BetaC = -(float(F.split('_')[2].lstrip('bn')))/100
                elif F.split('_')[2].lower().startswith('b'):
                    BetaC = (float(F.split('_')[2].lstrip('b').rstrip('.dat'))/100)

                CaseNumeric.append([MachC,AlphaC,BetaC])

#          accumulator={}#holds all the values of each variable together for everyfile
#          Zone_Accumulator= {}#holds  the values of each variable per zone and file
          Q=[]
          Flight_params=[]
          big_DataFrame=pd.DataFrame()

          COUNT=[]
          for  Ind, F in enumerate(Txtfiles):#Check the flight parameters and compare with the FILE to append q if they are in.
              Mach = (float(F.split('_')[2].lstrip('M')))
              Alpha = (float(F.split('_')[3].lstrip('a')))
              Beta = (float(F.split('_')[4].lstrip('b').rstrip('.dat')))
              Flight_params.append([Mach, Alpha, Beta])

          for Fpos,Fval in enumerate(Flight_params):
               for P,C in enumerate(CaseNumeric):
                  if Fval == C:
                     Q.append(Q_excel_list[P])
                     break
               else:
                     COUNT.append(Fpos)

          COUNT.sort(reverse=True) #sorted to start deleting from the end, not to modify the order of the  elements when deleting

          for i in COUNT:
               del Flight_params[i]
               del Txtfiles[i]


          print('file_cases not present on Excel: ','\n',list(set(Txtfiles2)-set(Txtfiles)),'\n')
          print('TOTAL nªFiles: ',len(Txtfiles)+len(COUNT),'\nNº missing files: ',len(COUNT))

          [Flight_params[qpos].append(qval) for qpos, qval in enumerate(Q)] #TO EVERY POSITION OF THE FLIGHT PARAMETER appends the CORRESPONDING vALUE
          [ CP_0, X_0, Y_0, Z_0, Array ] = Txt_Reader(Txtfiles[0])

          for  Ind, F in enumerate(Txtfiles):
              [ CP, X, Y, Z, Array ] = Txt_Reader(F)

#              accumulator[F] = {'cp':CP ,'x':X_0 ,'y':Y_0 ,'z':Z_0 }#,'Mach':mach, 'alpha':alpha, 'Beta':beta }
#              Zone_Accumulator[F] = Array

              ZippedList =  list(zip(CP, X_0, Y_0, Z_0))#,mach,alpha,beta))
              Inter = pd.DataFrame( ZippedList, columns=['cp','x','y','z'])#,'Mach','alpha','beta'])
              Inter['Mach'], Inter['Alpha'], Inter['Beta'], Inter['q'] = Flight_params[Ind]
              big_DataFrame = big_DataFrame.append(Inter)


     #     CPint=[]#attempt to create  a DF with the zones specified
     #     Xint=[]
     #     Yint=[]
     #     Zint=[]
     #     for file,zones in Zone_Accumulator.items():
     #         for zon,dims in zones.items():
     #              for dim,val in dims.items():
     #                          if dim=='cp':
     #                               CPint.extend(val)
     #                          elif dim=='x':
     #                               Xint.extend(val)
     #                          elif dim=='y':
     #                               Yint.extend(val)
     #                          elif dim=='z':
     #                               Zint.extend(val)
     #     ZippedList2 =  list(zip(CPint,Xint,Yint,Zint))
     #     Inter2=pd.DataFrame(ZippedList2,])


          Time1=time.time()-START
          print('time reading files: ',round(Time1,2),'seconds \n')
          df=big_DataFrame.copy()
          df=df.reset_index(drop=True)
          df=df.sort_values(by=['y'], inplace=False)
          df.to_csv('CSV350' ,index=None)#write the file 27 seconds



#
     return df
#     return df1,  Txtfiles , big_DataFrame, Flight_params,accumulator#, , Zone_Accumulator #,accumulator,Zone_Accumulator,Inter2,Flight_params,



df = DataFrame_Creator()

df = df[df.y != 0]#deletes the row if the value of y ==0
df = df.reset_index(drop=True)

range_Mach = [max( df ['Mach'] ),min( df['Mach'] )]
range_Alpha = [max( df ['Alpha'] ),min( df ['Alpha'] )]



#######

def calculate_fullmodel_by_points(poly,model1):# the first definition of the model(line 334) is commented in order to feed the function in a loop  with a set of variable parameters

     def modelo_punto( INTER, model1):
          X   = INTER[['Alpha','Mach']]
          Y   = INTER[['cp']].values

          scaler = StandardScaler()
          X1 = scaler.fit_transform(X)
          Mean = scaler.mean_
          Std_dev = scaler.scale_

          X_train, X_test, cp_train, cp_test= train_test_split( X1 ,Y,  test_size=0.2 ,random_state=5)#division of the train/test set of points


          X_poly_train = poly.fit_transform(X_train)
          X_poly_test = poly.fit_transform(X_test)

#
          cp_test=cp_test.squeeze()
          cp_train=cp_train.squeeze()



     #####BAYESIAN_RIDGE
     #############lambda--> precision of the weights
#     #############ALPHA--> precision of the NOISE

#          model1 = linear_model.BayesianRidge(####implement gridSearch to optimize parameteres
#                                    n_iter=100000,
#                                    tol=1e-6,
#                                    alpha_1=10000,#shAPE
#                                    alpha_2=1000,#RATE
#                                    lambda_1=500, #SHAPE
#                                    lambda_2=5000,#RATE
#                                    compute_score=False,
#                                    fit_intercept=True,
#                                    normalize=False,
#                                    copy_X=True,#not change
#                                    verbose=False)#centra el peso de todos los coeficientes alrededr de 0--> mas estable



          model1.fit( X_poly_train, cp_train )


          train_score = model1.score( X_poly_train, cp_train)
          test_score = model1.score( X_poly_test, cp_test)


          coeff = pd.DataFrame(model1.coef_, columns=['Coefficients Metamodel'])#used for lasso without polynomil transform


          return  train_score,  test_score, coeff, Mean, Std_dev

     TRAIN_SCORE  = []
     TEST_SCORE   = []
     COEFFICIENTS = []


     Inter = pd.DataFrame()
     Shape = df[['x', 'y','z']].drop_duplicates().dropna()
     contador = []

     for p, val_y in enumerate(Shape['y']):

          Inter = df[ df['y'] == val_y ]

          [   train_score, test_score, coeff, Mean, Std_dev ] = modelo_punto(  Inter, model1 )

#          if test_score>0:
#               CP.append( np.asscalar( cp_predicted ))
          TRAIN_SCORE.append( train_score )
          TEST_SCORE.append( test_score )

          COEFFICIENTS.append(coeff.values.T )


#          else:
          contador.append(p)

     Shape=Shape.reset_index( drop=True )
     Shape.drop( contador, inplace=True )

     print('MEAN_TEST :' , sum( TEST_SCORE  ) / len( TEST_SCORE ))
     print('MEAN_TRAIN:' , sum( TRAIN_SCORE ) / len( TRAIN_SCORE ))


     return TRAIN_SCORE, TEST_SCORE, COEFFICIENTS, Shape, Mean, Std_dev



def Coef_writer(path3, path2, path1, Params, pol, model1):#this function is implemented for speed, creating a model takes around 4 min, and load the specified variables from a file is much faster. not placed inside the "calculate_fullmodel_by_points" function in case just one hyperparameters set wants to be tried without placing it in a loop
     poly = PolynomialFeatures(pol,include_bias=True)
     if  os.path.isfile(os.path.join(path3,"Coefficients.npy")):
#           pass
           COEFFICIENTS = np.load(os.path.join(path3,"Coefficients.npy"))

           Shape = pd.DataFrame(np.load(os.path.join(path3,'Shape.npy')), columns=['x','y','z'])
           MEAN = np.load(os.path.join(path3,'Mean.npy'))
           STD_DeV = np.load(os.path.join(path3,'Standard_Deviation.npy'))



     else:
         try:
             os.makedirs(path3)

         except FileExistsError:
              pass

         finally:
          [ TRAIN_SCORE,  TEST_SCORE,  COEFFICIENTS, Shape, MEAN, STD_DeV ] = calculate_fullmodel_by_points(poly, model1)#, contador
          np.save(os.path.join(path3,"Coefficients"), COEFFICIENTS)
          np.save(os.path.join(path3,"Mean"), MEAN)
          np.save(os.path.join(path3,"Standard_Deviation"), STD_DeV)
          np.save(os.path.join(path3,'Shape'),Shape)
#          if   not os.path.isfile( 'R2_Scores.txt')
          with open(os.path.join(path2,'R2_Scores.txt'),'a') as file:#'a' stands for append
               file.write(Params+'\n')
               file.write('train score: '+ str(sum(TRAIN_SCORE)/len(TRAIN_SCORE))+'\n')
               file.write('test score: '+ str(sum(TEST_SCORE)/len(TEST_SCORE))+'\n')


     return Shape, COEFFICIENTS, MEAN, STD_DeV


# ####################################     System modelization    ##################################

def DATABASE_creator():# this function uses multiple loops to vary the hyperparameters
    parent_dir = r'Z:\ba_NG8569F_Pedro\Belly_Fairing\01_BF_Data\A350'

#     Modelic=str(input('introduce the name of the model you want to use: BayesionRidge(1)/LinearRegression(2):'))
    Modelic='BayesianRidge'

    if Modelic == 'BayesianRidge' :

         def Excel_Trazability_BR():
               f = open(path2+'\R2_scores.txt','r')
               Lines = f.readlines()
               f.close()
               Train_Scores = []
               Test_Scores =[]# np.zeros(int(len(Lines)/3))
               A1 = []
               A2 = []
               L2 = []
               L1 = []
               tol = []

               Train_Scores.extend( float(i.split(': ')[1]) for i in Lines if i.startswith('train score:'))
               Test_Scores.extend( float(i.split(': ')[1] )for i in Lines if i.startswith('test score:'))
               A1.extend( float(i.split(',')[0].split('=')[1])for i in Lines if i.startswith('alpha'))
               A2.extend( float(i.split(',')[1].split('=')[1])for i in Lines if i.startswith('alpha'))
               L1.extend( float(i.split(',')[2].split('=')[1])for i in Lines if i.startswith('alpha'))
               L2.extend( float(i.split(',')[3].split('=')[1])for i in Lines if i.startswith('alpha'))
               tol.extend( float(i.split(',')[5].split('=')[1])for i in Lines if i.startswith('alpha'))
               ZippedList =  list(zip(A1, A2, L1, L2, tol, Train_Scores, Test_Scores))
               Table = pd.DataFrame( ZippedList, columns=['alpha1','alpha2','lambda1','lambda2', 'tolerance','Train Score','Test Score' ])


               with pd.ExcelWriter(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), engine="openpyxl", mode='a') as writer:
                    Table.to_excel(writer,sheet_name='order'+f"{pol}")

         for pol in [2]:
     #     pol = 3,,
              poly = PolynomialFeatures(pol,include_bias=True)
              A1=[ 1e-5, 1e-2, 1e2, 1e7, 0.1, 1, 5, 10, 500]
              A1.reverse()
#              A2=[0.001]#different variables can be assigned for each loop
              for a1 in A1:
                   for a2 in [ 1e-5, 1e-2, 1e2, 1e7, 0.1, 1, 5, 10, 500]:
                        for l1 in A1:
                             for l2 in A1:



                              model1 = BayesianRidge(n_iter=1000,
                                                        tol=1e-2,
                                                        alpha_1=a1,#shAPE
                                                        alpha_2=a2,#RATE
                                                        lambda_1=l1, #SHAPE
                                                        lambda_2=l2)#RATE


                              Method = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[0]# presented this way to be able to change the method--> implrment an if condition for the parameters of other methods.
                              Params = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[1].replace('\n','').replace(' ','').replace(')>','').replace('compute_score=False','').replace('copy_X=True,','').replace('fit_intercept=True,','').replace('verbose=False','').replace('normalize=False,','').replace(',,',',').rstrip(",")
                              path1 = os.path.join(parent_dir, Method)
                              path2 = os.path.join(path1, 'order'+str(pol))
                              path3 = os.path.join(path2, Params)

                              [Shape, COEFFICIENTS, MEAN, STD_DeV ] = Coef_writer(path3, path2, path1, Params, pol, model1)

              try:
                   wb = load_workbook(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), read_only=False)

              except FileNotFoundError:
                    wb = openpyxl.Workbook()
                    wb.save(os.path.join(path1, 'Trazabilidad350PIP.xlsx'))

#              if not 'order'+f"{pol}" in wb.sheetnames:
              Excel_Trazability_BR()


    elif Modelic == 'ElasticNetCV' :

         def Excel_Trazability_ElastNet():
               f = open(path2+'\R2_scores.txt','r')
               Lines = f.readlines()
               f.close()
               Train_Scores = []
               Test_Scores =[]# np.zeros(int(len(Lines)/3))
               A1 = []
               A2 = []
               L2 = []
               L1 = []
               tol = []

               Train_Scores.extend( float(i.split(': ')[1]) for i in Lines if i.startswith('train score:'))
               Test_Scores.extend( float(i.split(': ')[1] )for i in Lines if i.startswith('test score:'))
               A1.extend( float(i.split(',')[0].split('=')[1])for i in Lines if i.startswith('alpha'))
               A2.extend( float(i.split(',')[1].split('=')[1])for i in Lines if i.startswith('alpha'))
               L1.extend( float(i.split(',')[2].split('=')[1])for i in Lines if i.startswith('alpha'))
               L2.extend( float(i.split(',')[3].split('=')[1])for i in Lines if i.startswith('alpha'))
               tol.extend( float(i.split(',')[5].split('=')[1])for i in Lines if i.startswith('alpha'))
               ZippedList =  list(zip(A1, A2, L1, L2, tol, Train_Scores, Test_Scores))
               Table = pd.DataFrame( ZippedList, columns=['alpha1','alpha2','lambda1','lambda2', 'tolerance','Train Score','Test Score' ])


               with pd.ExcelWriter(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), engine="openpyxl", mode='a') as writer:
                    Table.to_excel(writer,sheet_name='order'+f"{pol}")

         for pol in [3, 4]:
     #     pol = 3,,
              for e in  [0.001,0.1]:
                   poly = PolynomialFeatures(pol,include_bias=True)

                   model1 = linear_model.ElasticNetCV(
                        l1_ratio = [0.1, 0.5, 0.99],
                        eps=e,
                        tol=1e-2,
                        selection='random',
                        precompute='auto',
                        max_iter=1e5,
                        fit_intercept=False,
                        normalize=False,
                        n_jobs=-1,
                        cv=2,
                        n_alphas=10)


#l1_ratio=0.5, eps=0.001, n_alphas=100, alphas=None, fit_intercept=True, normalize=False, precompute=’auto’, max_iter=1000, tol=0.0001, cv=’warn’, copy_X=True, verbose=0, n_jobs=None, positive=False, random_state=None, selection=’cyclic’)

                   Params = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[1].replace('\n','').replace(' ','').replace(')>','').replace('compute_score=False','').replace('copy_X=True,','').replace('fit_intercept=True,',''). replace('verbose=False','').replace('normalize=False,','').replace(',,',',').rstrip(",")

                   path1 = os.path.join(parent_dir, Modelic)
                   path2 = os.path.join(path1, 'order'+str(pol))
                   path3 = os.path.join(path2, Params)

                   [Shape, COEFFICIENTS, MEAN, STD_DeV ] = Coef_writer(path3, path2, path1, Params, pol, model1)

              try:
                   wb = load_workbook(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), read_only=False)

              except FileNotFoundError:
                    wb = openpyxl.Workbook()
                    wb.save(os.path.join(path1, 'Trazabilidad350PIP.xlsx'))

#              if not 'order'+f"{pol}" in wb.sheetnames:
              Excel_Trazability_ElastNet()




    elif Modelic == 'LinearRegression' :

        def Excel_Trazability_LR():

               f = open(path2+'\R2_scores.txt','r')
               Lines = f.readlines()
               f.close()
               Train_Scores = []
               Test_Scores =[]# np.zeros(int(len(Lines)/3))
               fit = []
               nor = []
               cop = []

               Train_Scores.extend( float(i.split(': ')[1]) for i in Lines if i.startswith('train score:'))
               Test_Scores.extend( float(i.split(': ')[1] )for i in Lines if i.startswith('test score:'))
               fit.extend( i.split(',')[1].split('=')[1]for i in Lines if i.startswith('copy'))
               nor.extend(i.split(',')[2].split('=')[1]for i in Lines if i.startswith('copy'))
               cop.extend( i.split(',')[0].split('=')[1]for i in Lines if i.startswith('copy'))

               ZippedList =  list(zip(fit,nor,cop, Train_Scores, Test_Scores))
               Table = pd.DataFrame( ZippedList, columns=['fit_intercept','normalize','copy_X','Train Score','Test Score' ])

               with pd.ExcelWriter(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), engine="openpyxl", mode='a') as writer:#extremadamente recomendable usar "with"--> cierra los excel al terminar de usarlos
                    Table.to_excel(writer,sheet_name='order'+f"{pol}")

        Method = Modelic
        path1 = os.path.join(parent_dir, Method)

        try:#carga o crea el archivo de trazabilidades
                wb = load_workbook(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), read_only=False)
        except FileNotFoundError:
                    wb = openpyxl.Workbook()
                    wb.save(os.path.join(path1, 'Trazabilidad350PIP.xlsx'))

        for pol in [2,3,4,5]:
             for f in [False , True]:
                   for n in [False , True]:
                        for c in [False , True]:
                             poly = PolynomialFeatures(pol,include_bias=True)
                             model1=LR( n_jobs = -1,
                                        fit_intercept = f,
                                        normalize = n,
                                        copy_X = c)




                             Params = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[1].replace('\n','').replace(' ','').replace(')>','').rstrip(",").replace('n_jobs=-1','').replace(',,',',')


                             path2 = os.path.join(path1, 'order'+str(pol))
                             path3 = os.path.join(path2, Params)

                             [Shape, COEFFICIENTS, MEAN, STD_DeV ] = Coef_writer(path3, path2, path1, Params, pol, model1)

             try:#carga o crea el archivo de trazabilidades
                     wb = load_workbook(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), read_only=False)
             except FileNotFoundError:
                         wb = openpyxl.Workbook()
                         wb.save(os.path.join(path1, 'Trazabilidad350PIP.xlsx'))

             if not 'order'+f"{pol}" in wb.sheetnames:
                     Excel_Trazability_LR()



    elif Modelic == 'LassoLarsCV' :


        def Excel_Trazability_LassoLarsCV():

               f = open(path2+'\R2_scores.txt','r')
               Lines = f.readlines()
               f.close()
               Train_Scores = []
               Test_Scores =[]# np.zeros(int(len(Lines)/3))
               eps = []
               cv = []
               n_a = []

               Train_Scores.extend( float(i.split(': ')[1]) for i in Lines if i.startswith('train score:'))
               Test_Scores.extend( float(i.split(': ')[1] )for i in Lines if i.startswith('test score:'))
               eps.extend( i.split(',')[2].split('=')[1]for i in Lines if i.startswith('copy'))
               cv.extend(i.split(',')[1].split('=')[1]for i in Lines if i.startswith('copy'))
               n_a.extend( i.split(',')[3].split('=')[1]for i in Lines if i.startswith('copy'))

               ZippedList =  list(zip(eps,cv,n_a, Train_Scores, Test_Scores))
               Table = pd.DataFrame( ZippedList, columns=['fit_intercept','normalize','copy_X','Train Score','Test Score' ])

               with pd.ExcelWriter(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), engine="openpyxl", mode='a') as writer:#extremadamente recomendable usar "with"--> cierra los excel al terminar de usarlos
                    Table.to_excel(writer,sheet_name='order'+f"{pol}")


        P = [3,4]
#        P.reverse()
        E = [1e-20, 1e-10, 0, 1000]
#        E.reverse()

#        CV.reverse()
        A = [ 1, 5, 10,100]
        for pol in P:
             for e in E:
                  for a in A:

                             poly = PolynomialFeatures(pol,include_bias=True)
                             model1 = LassoCV(cv=3 ,
                                         verbose=False,
                                         normalize=False,
                                         eps=e,
                                         max_iter=1e9,
                                         n_jobs=-1,
                                         copy_X=True,
                                         random_state=55,
                                         selection='random',
                                         tol=0.01,
                                         n_alphas=a)#n_alphas needs to be an integer

                             Method = Modelic

                             Params = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[1].replace('\n','').replace(')>','').replace('compute_score=False','').replace('copy_X=True,','').replace('fit_intercept=True,','').replace('verbose=False','').replace('normalize=False,','').replace(',,',',').replace(', n_jobs=-1,','')             .replace('positive=False','').replace(" precompute='auto',",'').replace(',             max_iter=1000000.0','').replace(' ','').rstrip(",")

                             path1 = os.path.join(parent_dir, Method)
                             path2 = os.path.join(path1, 'order'+str(pol))
                             path3 = os.path.join(path2, Params)

                             [Shape, COEFFICIENTS, MEAN, STD_DeV ] = Coef_writer(path3, path2, path1, Params, pol, model1)
             try:#carga o crea el archivo de trazabilidades
                     wb = load_workbook(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), read_only=False)
             except FileNotFoundError:
                         wb = openpyxl.Workbook()
                         wb.save(os.path.join(path1, 'Trazabilidad350PIP.xlsx'))
                         wb.remove_sheet('sheet')


             if not 'order'+f"{pol}" in wb.sheetnames:
                     Excel_Trazability_LassoLarsCV()


    elif  Modelic == 'LassoCV' :


        def Excel_Trazability_LassoCV():

               f = open(path2+'\R2_scores.txt','r')
               Lines = f.readlines()
               f.close()
               Train_Scores = []
               Test_Scores =[]# np.zeros(int(len(Lines)/3))
               fit = []
               nor = []
               cop = []

               Train_Scores.extend( float(i.split(': ')[1]) for i in Lines if i.startswith('train score:'))
               Test_Scores.extend( float(i.split(': ')[1] )for i in Lines if i.startswith('test score:'))
               fit.extend( i.split(',')[1].split('=')[1]for i in Lines if i.startswith('copy'))
               nor.extend(i.split(',')[2].split('=')[1]for i in Lines if i.startswith('copy'))
               cop.extend( i.split(',')[0].split('=')[1]for i in Lines if i.startswith('copy'))

               ZippedList =  list(zip(fit,nor,cop, Train_Scores, Test_Scores))
               Table = pd.DataFrame( ZippedList, columns=['fit_intercept','normalize','copy_X','Train Score','Test Score' ])

               with pd.ExcelWriter(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), engine="openpyxl", mode='a') as writer:#extremadamente recomendable usar "with"--> cierra los excel al terminar de usarlos
                    Table.to_excel(writer,sheet_name='order'+f"{pol}")

        P = [2,3,4]
#        P.reverse()
        E = [1e-20, 1e-10, 0, 1000]
#        E.reverse()
        CV = [2,3,4]
#        CV.reverse()
        A = [ 1, 5, 10,100]
        for pol in P:
             for e in E:
                   for cv in CV:
                        for a in A:

                             poly = PolynomialFeatures(pol,include_bias=True)
                             model1 = LassoCV(cv=cv ,
                                         verbose=False,
                                         normalize=False,
                                         eps=e,
                                         max_iter=1e9,
                                         n_jobs=-1,
                                         copy_X=True,
                                         selection='random',
                                         tol=0.01,
                                         n_alphas=a)#n_alphas needs to be an integer

                             Method = Modelic

                             Params = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[1].replace('\n','').replace(')>','').replace('compute_score=False','').replace('copy_X=True,','').replace('fit_intercept=True,','').replace('verbose=False','').replace('normalize=False,','').replace(',,',',').replace(', n_jobs=-1,','')             .replace('positive=False','').replace(" precompute='auto',",'').replace(',             max_iter=1000000.0','').replace(' ','').rstrip(",")

                             path1 = os.path.join(parent_dir, Method)
                             path2 = os.path.join(path1, 'order'+str(pol))
                             path3 = os.path.join(path2, Params)

                             [Shape, COEFFICIENTS, MEAN, STD_DeV ] = Coef_writer(path3, path2, path1, Params, pol, model1)
             try:#carga o crea el archivo de trazabilidades
                     wb = load_workbook(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), read_only=False)
             except FileNotFoundError:
                         wb = openpyxl.Workbook()
                         wb.save(os.path.join(path1, 'Trazabilidad350PIP.xlsx'))
             if not 'order'+f"{pol}" in wb.sheetnames:
                     Excel_Trazability_LassoCV()




    return Shape , COEFFICIENTS, MEAN, STD_DeV, poly, path3, path2, path1

[Shape , COEFFICIENTS, MEAN, STD_DeV, poly, path3, path2, path1] = DATABASE_creator()





END7=time.time()
FULLTIME=END7-START
print('full time : ', round(FULLTIME,2), 'seconds')



"""

#333
#
#
#
##
##
############################# Selection of the model parameters for the GUI
##-think about implementing a Excel search for maximum test values
parent_dir = r'Z:\ba_NG8569F_Pedro\Belly_Fairing\01_BF_Data\A350'
dictt = dict(max_n_alphas=1000, n_jobs=-1, eps=2.220446049250313e-16)

model1 = svc = svm.SVC(C=1, kernel='linear')
model1.set_params(dictt)
pol=5
poly = PolynomialFeatures(pol,include_bias=True)

Method = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[0]# presented this way to be able to change the method--> implrment an if condition for the parameters of other methods.
Params = str(model1.get_params).replace('<bound method BaseEstimator.get_params of ','').split('(')[1].replace('\n','').replace(' ','').replace(')>','').replace('compute_score=False','').replace('copy_X=True,','').replace('fit_intercept=True,','').replace('verbose=False','').replace('normalize=False,','').replace(',,',',').rstrip(",")
path1 = os.path.join(parent_dir, Method)
path2 = os.path.join(path1, 'order'+str(pol))
path3 = os.path.join(path2, Params)

[Shape, COEFFICIENTS, MEAN, STD_DeV ] = Coef_writer(path3, path2, path1, Params, pol, model1)











#
##
################################################  TKinter (GUI)   ###############################################

#from matplotlib.backends.backend_tkagg import FigureCAnvasTkAgg, NavigationToolbar2Tk

root = Tk()

root.title("Cp_e vs Flight parameters")
root.iconbitmap(r'Z:\ba_NG8569F_Pedro\Belly_Fairing\GUI\airbus-icon-3.ico')





tab_parent = ttk.Notebook(root)
tab1 = ttk.Frame(tab_parent)
tab2 = ttk.Frame(tab_parent)

tab_parent.add(tab1, text="    350    ",padding=5 )
tab_parent.add(tab2, text="    380    ",padding=5)
tab_parent.pack(expand=1, fill='both')

tab_parent.enable_traversal()

my_image = ImageTk.PhotoImage(Image.open(r'Z:\ba_NG8569F_Pedro\Belly_Fairing\GUI\BF.jpg'))

my_label = Label(master=tab1, image = my_image , relief='sunken', bd=3)
my_label.grid(row=0, column=1)

def destroyy():
     root.quit()
     root.destroy()

button_destroy = Button( tab1, text='Exit',  command = destroyy)#not working
button_destroy.grid( row=0, column=0, padx=50  )#,columnspan=1



COEFFICIENTS = np.load(os.path.join(path3,"Coefficients.npy"))
Shape = pd.DataFrame(np.load(os.path.join(path3,'Shape.npy')), columns=['x','y','z'])
MEAN = np.load(os.path.join(path3,'Mean.npy'))
STD_DeV = np.load(os.path.join(path3,'Standard_Deviation.npy'))



def PlotTKinter(var):# pasa un valor de variable al usar un slider

     Mach=mach.get()
     Alpha=alpha.get()


     CP_new  = np.zeros(Shape.shape[0])

     desired_params =pd.DataFrame([( Alpha, Mach)], columns=[ 'Alpha', 'Mach' ])
     desired_std_params = (desired_params.sub(MEAN)).div(STD_DeV)
     Params_poly= poly.fit_transform( desired_std_params )

     for pos, data in enumerate(COEFFICIENTS):
          CP_new[pos]=np.dot(np.array(data),np.array(Params_poly).T)

     figure1 = plt.Figure(figsize=(8,6), dpi=100,)
     ax = figure1.add_subplot(111, projection='3d')

     x = Shape['x']
     y = Shape['y']
     z = Shape['z']
     cs = CP_new
     title = "Alpha={} Mach={}" .format (Alpha,Mach)



     cm = plt.get_cmap('rainbow')
     cNorm = matplotlib.colors.Normalize(vmin=min(cs), vmax=max(cs))
     scalarMap = cmx.ScalarMappable(norm=cNorm, cmap=cm)


     scalarMap.set_array(cs)
     figure1.colorbar(scalarMap).set_label('Cp', rotation=90, fontsize=14, fontweight='bold')


     ax.set_xlabel('x',fontsize=14, fontweight='bold')
     ax.set_ylabel('y',fontsize=14, fontweight='bold')
     ax.set_zlabel('z',fontsize=14, fontweight='bold')
     ax.set_title(title,fontsize=10,fontweight='bold')
#     ax.legend()



#     ax.scatter(x, y, z, c=scalarMap.to_rgba(cs))

     max_range = np.array([max(x)-min(x), max(y)-min(y), max(z)-min(z)]).max() / 2.0
     mid_x = (max(x)+min(x)) * 0.5
     mid_y = (max(y)+min(y)) * 0.5
     mid_z = (max(z)+min(z)) * 0.5
     ax.set_xlim(mid_x - max_range, mid_x + max_range)
     ax.set_ylim(mid_y - max_range, mid_y + max_range)
     ax.set_zlim(mid_z - max_range, mid_z + max_range)

     ax.scatter(x, y, z, c=scalarMap.to_rgba(cs))#plotea en los ejes

     canvas = FigureCanvasTkAgg( figure1, master=tab1,)  # crea un marco
     canvas.get_tk_widget().grid( row=0, column=1,)#paints the image



     FrameNav = Frame(tab1,width=5, height=2, relief='raised', )# toolbar solo puede ponerse en la main window con "pack" , por eso creo un Frame para "engañar" al programa y ponerlo donde quiera
     FrameNav.grid (row=1, column=1, sticky = 'sw')

     FrameNav1 = Canvas(FrameNav,width=355, height=2, )
     FrameNav1.grid (row=0, column=0)


     FrameNav2 = Frame(FrameNav,width=55, height=2, relief='raised', borderwidth=5)
     FrameNav2.grid (row=0, column=1)



     toolbar = NavigationToolbar2Tk(canvas, FrameNav2)#toma los datos de "canvas" y los pinta en "FrameNav"
     toolbar.update()#configurar los botones alante y atras pa ver las graficas acumuladas



     ax.mouse_init()#veryimnportant to play the mouse_init after setting the canvas of the main window





frameA = Frame(tab1)
frameA.grid(row=0, column=2, pady=0, )

alpha = IntVar()
Mach= IntVar()

alpha = Scale(frameA, from_= range_Alpha[1], to = range_Alpha[0], orient= 'horizontal', digits=4, label='Alpha',resolution=0.05, variable= alpha, cursor = 'hand2 ',command = PlotTKinter)
alpha.pack(fill = 'both', expand = True,side= 'bottom')



mach = Scale(frameA, from_= range_Mach[1], to = range_Mach[0], orient= 'horizontal',digits=2, label='Mach', resolution=0.05,variable= Mach, cursor = 'hand2 ' ,command = PlotTKinter)
mach.pack(fill = 'both', expand = True,  side= 'bottom')


#
#my_btn = Button(frameA, text="Plot!", command = PlotTKinter)
#my_btn.pack(fill = 'x', expand = False,)


root.quit()
root.mainloop()

#def destroyy():
#     root.quit()
#     root.destroy()
#
#button_destroy = Button( root, text='Exit',  command = destroyy)#not working
#button_destroy.grid( row=0, column=0, padx=50  )#,columnspan=1
#
#
#
#COEFFICIENTS = np.load(os.path.join(path3,"Coefficients.npy"))
#Shape = pd.DataFrame(np.load(os.path.join(path3,'Shape.npy')), columns=['x','y','z'])
#MEAN = np.load(os.path.join(path3,'Mean.npy'))
#STD_DeV = np.load(os.path.join(path3,'Standard_Deviation.npy'))
#
#
#
#def PlotTKinter(var):# pasa un valor de variable al usar un slider
#
#     Mach=mach.get()
#     Alpha=alpha.get()
#
#
#     CP_new  = np.zeros(Shape.shape[0])
#
#     desired_params =pd.DataFrame([( Alpha, Mach)], columns=[ 'Alpha', 'Mach' ])
#     desired_std_params = (desired_params.sub(MEAN)).div(STD_DeV)
#     Params_poly= poly.fit_transform( desired_std_params )
#
#     for pos, data in enumerate(COEFFICIENTS):
#          CP_new[pos]=np.dot(np.array(data),np.array(Params_poly).T)
#
#     figure1 = plt.Figure(figsize=(8,6), dpi=100,)
#     ax = figure1.add_subplot(111, projection='3d')
#
#     x = Shape['x']
#     y = Shape['y']
#     z = Shape['z']
#     cs = CP_new
#     title = "Alpha={} Mach={}" .format (Alpha,Mach)
#
#
#
#     cm = plt.get_cmap('rainbow')
#     cNorm = matplotlib.colors.Normalize(vmin=min(cs), vmax=max(cs))
#     scalarMap = cmx.ScalarMappable(norm=cNorm, cmap=cm)
#
#
#     scalarMap.set_array(cs)
#     figure1.colorbar(scalarMap).set_label('Cp', rotation=90, fontsize=14, fontweight='bold')
#
#
#     ax.set_xlabel('x',fontsize=14, fontweight='bold')
#     ax.set_ylabel('y',fontsize=14, fontweight='bold')
#     ax.set_zlabel('z',fontsize=14, fontweight='bold')
#     ax.set_title(title,fontsize=10,fontweight='bold')
##     ax.legend()
#
#
#
##     ax.scatter(x, y, z, c=scalarMap.to_rgba(cs))
#
#     max_range = np.array([max(x)-min(x), max(y)-min(y), max(z)-min(z)]).max() / 2.0
#     mid_x = (max(x)+min(x)) * 0.5
#     mid_y = (max(y)+min(y)) * 0.5
#     mid_z = (max(z)+min(z)) * 0.5
#     ax.set_xlim(mid_x - max_range, mid_x + max_range)
#     ax.set_ylim(mid_y - max_range, mid_y + max_range)
#     ax.set_zlim(mid_z - max_range, mid_z + max_range)
#
#     ax.scatter(x, y, z, c=scalarMap.to_rgba(cs))#plotea en los ejes
#
#     canvas = FigureCanvasTkAgg( figure1, master=root,)  # crea un marco
#     canvas.get_tk_widget().grid( row=0, column=1,)#paints the image
#
#
#
#     FrameNav = Frame(root,width=5, height=2, relief='raised', )# toolbar solo puede ponerse en la main window con "pack" , por eso creo un Frame para "engañar" al programa y ponerlo donde quiera
#     FrameNav.grid (row=1, column=1, sticky = 'sw')
#
#     FrameNav1 = Canvas(FrameNav,width=355, height=2, )
#     FrameNav1.grid (row=0, column=0)
#
#
#     FrameNav2 = Frame(FrameNav,width=55, height=2, relief='raised', borderwidth=5)
#     FrameNav2.grid (row=0, column=1)
#
#
#
#     toolbar = NavigationToolbar2Tk(canvas, FrameNav2)#toma los datos de "canvas" y los pinta en "FrameNav"
#     toolbar.update()#configurar los botones alante y atras pa ver las graficas acumuladas
#
#
#
#     ax.mouse_init()#veryimnportant to play the mouse_init after setting the canvas of the main window
#
#
#
#
#
#frameA = Frame(root)
#frameA.grid(row=0, column=2, pady=0, )
#
#alpha = IntVar()
#Mach= IntVar()
#
#alpha = Scale(frameA, from_= range_Alpha[1], to = range_Alpha[0], orient= 'horizontal', digits=4, label='Alpha',resolution=0.05, variable= alpha, cursor = 'hand2 ',command = PlotTKinter)
#alpha.pack(fill = 'both', expand = True,side= 'bottom')
#
#
#
#mach = Scale(frameA, from_= range_Mach[1], to = range_Mach[0], orient= 'horizontal',digits=2, label='Mach', resolution=0.05,variable= Mach, cursor = 'hand2 ' ,command = PlotTKinter)
#mach.pack(fill = 'both', expand = True,  side= 'bottom')
#
#
##
##my_btn = Button(frameA, text="Plot!", command = PlotTKinter)
##my_btn.pack(fill = 'x', expand = False,)
#
#
#root.quit()
#root.mainloop()

##
#


#
#
#
#
#
#def modelito(DATA,pol,title,Sections):
#     tt=time.time()
##     End1=time.time()
#     #     X   = df[['q','Beta','Alpha','Mach','x','y','z']]
##          Y   = df['cp']
#     X   = DATA[['q','Alpha','Mach','x','y','z']]#'x','y','z'#selection of multiple columns in dataframe needs [[ ]]
#     Y   = DATA[['cp']].values# target function needs to be a Pandas Series/array
#     #     https://scikit-learn.org/stable/auto_examples/model_selection/plot_underfitting_overfitting.html#sphx-glr-auto-examples-model-selection-plot-underfitting-overfitting-py
#
#          ##     #apply StandardScaler# ¿is it necesary ¿¿for X and cp??? or fit_transform already standarizes???
##     scaler=StandardScaler()
##     X1=scaler.fit_transform(X)#keeps the original columns of the DF for the names of the parameters
#     X_train, X_test, cp_train, cp_test= train_test_split( X ,Y,  test_size=0.2 ,random_state=5)#division of the train/test set of points
#
#     poly = PolynomialFeatures(pol,include_bias=False) #s-- Numerical value means order of the polinomial, try different orders to see lowest sensible model.
#     ##https://towardsdatascience.com/polynomial-regression-bbe8b9d97491
#     X_poly_train = poly.fit_transform(X_train)
##     print('size of XpolyTrain: ',X_poly_train.shape)
#     X_poly_test = poly.fit_transform(X_test)
#     X_poly_train_feature_name = poly.get_feature_names(list(X))
#
#
#     ##
#     #     #create dataframe from the polynomial X values and features
#     df_poly_train= pd.DataFrame(X_poly_train, columns=X_poly_train_feature_name)
#     #     #print(df_poly_train.head())
#
#     cp_test=cp_test.squeeze()
##     cp_train=scaler.fit_transform(cp_train)
#     cp_train=cp_train.squeeze()
#
#
#     df_poly_train['cp']=cp_train
#
##     classifiers = [
##         linear_model.SGDRegressor(),
##         linear_model.BayesianRidge(),
##         linear_model.LassoLars(),
##        linear_model.ARDRegression(),
##         linear_model.PassiveAggressiveRegressor(),
##         linear_model.TheilSenRegressor(),
##         linear_model.LinearRegression()]
##     for item in classifiers:
##         print(item)
##         clf = item
##         clf.fit(X_poly_train, cp_train)
##         print(clf.predict(X_poly_test),'\n')
##         print(clf.score)
#     ##
#     #     print('parameters of the model:' ,'\n',x_poly_train_feature_name,'\n')
#     #     print('number of parameters of the model: ',len(x_poly_train_feature_name))
#
##     End2=time.time()
##     Time2=End2-End1
##     print('time separating samples (test/train): ', Time2, 'seconds')
#
##     print('model used: Lasso_CV')
##     model1 = LassoCV(cv=3, verbose=False,
##                      normalize=False,
##                      eps=0.0000001,
##                      n_alphas=5,
##                      tol=1e-6,
##                      max_iter=1e9,
##                      n_jobs=-1,
##                      random_state=55 ,) #maxR^2=0.80
#
##     print('model used: LassoLARS_CV')
##     model1 = LassoLarsCV(cv=2 ,
##                          verbose=False,
##                          normalize=False,
##                          eps=1e-15,
##                          max_iter=1e6,
##                          n_jobs=-2,
##                          copy_X=True,
##                          max_n_alphas=1e5)#R^2=0.89 , coeff all equal to 0 minus x y z
#
##     print('model used: LarsCV')
##     model1 = LarsCV(cv=4 ,
##                     verbose=False,
##                     normalize=False,
##                     eps=1e-1,
##                     max_iter=1e6,
##                     n_jobs=-1,
##                     copy_X=True,
##                     max_n_alphas=1e5)#R^2=0.89
#
#     print('model used: Linear Regression')
#     model1=LR( n_jobs = -1,
#               fit_intercept = False,
#               normalize = False,
#               copy_X = True)
#
##     model1=LassoLarsIC(criterion='bic')
#
##     model1=LassoLarsIC(criterion='aic')#model r2=0.79, cuts 0.91
#
##     print('model used: ElasticNetCV')
##     model1 = ElasticNetCV( copy_X=True,
##                           fit_intercept=False,
##                           l1_ratio=[0.1,1,10],
##                           max_iter=1e6,
##                           cv=3,
##                           tol=1e-6,
##                           normalize=False)#
#
#
##     print('model used: Ridge_CV')
##     model1 = RidgeCV( alphas=(0.1, 1.0, 10.0,100.0),
##                    fit_intercept=False,
##                    normalize=True,
##                    scoring=None,
##                    cv=4,
##                    gcv_mode='svd')
#
##     print('model used: ARDRegression')
##     model1 = linear_model.ARDRegression()
#
##     model1 = linear_model.BayesianRidge(
##                                    n_iter=30,
##                                    tol=1e-3,
##                                    alpha_1=1e-3,
##                                    alpha_2=1e-3,
##                                    lambda_1=1e1,
##                                    lambda_2=1e11,
##                                    compute_score=True,
##                                    fit_intercept=False,
##                                    normalize=True,
##                                    copy_X=True,
##                                    verbose=False)#centra el peso de todos los coeficientes alrededr de 0--> mas estable
###
##     model1 = SGDRegressor( loss='squared_loss',
##                          penalty='l2',
##                          alpha=0.0001,
##                          l1_ratio=0.15,
##                          fit_intercept=False,
##                          max_iter=1000000000,
##                          tol=0.1,
##                          shuffle=False,
##                          verbose=0,
##                          epsilon=0.01,
##                          random_state=55,
##                          learning_rate='invscaling',
##                          eta0=0.01,
##                          power_t=0.25,
##                          early_stopping=False,
##                          validation_fraction=0.1,
##                          n_iter_no_change=5,
##                          warm_start=False,
##                          average=False)
#
##
##     model1 = LinearSVR( epsilon=1e-10000000,
##                        tol=1e-10,
##                        C=1e1,
##                        loss='squared_epsilon_insensitive',
##                        fit_intercept=True ,
##                        intercept_scaling=2e2,
##                        dual=False,
##                        random_state=55,
##                        max_iter=10000)#if dual=True ¿hangs..from 4s(and bad results)-->680s and not good results  //Dual= Flase....R2<0
#
#
#     model1.fit( X_poly_train, cp_train )
#
#
##     End4=time.time()
##     Time4=End4-End2
##     print('time fitting train data: ', Time4, 'seconds','\n')
#
##     # Predicted metamodel data:
#
##     cp_pred_train = np.array(model1.predict(X_poly_train))
##     cp_pred_test = np.array(model1.predict(X_poly_test))
##     print('cp_predicted (train): ',cp_pred_train )
##     print('cp_predicted(test): ',cp_pred_test)
##
##     End5=time.time()
##     Time5=End5-End4
##     print('time predicting data: ', Time5, 'seconds','\n')
#
##     RMSE_test = np.sqrt(np.sum(np.square(cp_test-cp_pred_test )))
##     RMSE_train = np.sqrt(np.sum(np.square(cp_train- cp_pred_train)))
##     RMSE_test  = np.sqrt(mse(cp_test,cp_pred_test ))
##     RMSE_train = np.sqrt(mse(cp_train, cp_pred_train))
#
##
##     print("Root mean square error of Metamodel(train):", RMSE_train,'\n')
##     print("Root mean square error of Metamodel(test):", RMSE_test,'\n')
##
#     train_score = model1.score( X_poly_train, cp_train)
#     test_score  = model1.score( X_poly_test,  cp_test)
#
##     print("Predicted Model R^2(train data):", train_score,'\n')
##     print("Predicted Model R^2(test data):", test_score,'\n')
##     #     coeff1=model1.coef_
#
#     coeff1 = pd.DataFrame( model1.coef_,index =df_poly_train.drop('cp',axis=1).columns, columns=['Coefficients Metamodel'])
#
#     cp_predicted = []
#
#     for pos, title in enumerate(title):
#
##          X_new1=scaler.fit_transform(Sections[pos])
#          X_new_poly = poly.fit_transform( Sections[pos] )
#          cp_predicted.append( model1.predict ( X_new_poly ))
#
#
#     # Check metamodel fit compared with test database:
#     #
#     #
#     #     # Response check Metamodel
#     #     #Typical dimension error: https://markhneedham.com/blog/2017/07/05/pandasscikit-learn-get_dummies-testtrain-sets-valueerror-shapes-not-aligned/
#     #
#     #     #create dataframe from the polynomial X values and features
#     df_poly_test = pd.DataFrame(X_poly_test, columns=X_poly_train_feature_name)
#
#     #Add response cp  to the new dataframe
#     df_poly_test['cp_test'] = cp_test
#
#
#
##         ### Plot predicted Vs test:
#
## ##########################################   5.-Plot/analyse results,       ##########################################
##
##     plt.figure(figsize=(12,8))
##     plt.xlabel("Predicted value with Metamodel",fontsize=20)
##     plt.ylabel("Actual y-values",fontsize=20)
##     plt.grid(1)
##     plt.scatter(cp_pred_train,cp_train,edgecolors=(0,0,0),lw=2,s=80)
##     plt.plot(cp_pred_train,cp_pred_train, 'k--', lw=2)
##     plt.title(str(title))
##     plt.show()
#
##
#     end=time.time()
#     TIME=end-tt
#
#     return coeff1, test_score, train_score, TIME ,cp_predicted#,range
#
#
##     ##################################################################################################################
#
#def sections_y():
#     start = time.time()
#     poly_order = 3#max 4
#     number_of_SECTIONS = 3#max 18 sections
#
#     #     while True:
#     try :
#          if number_of_SECTIONS<=0  :
#             number_of_SECTIONS = input( 'negative number of sections, introduce a positive value: ' )
#
#
#     except TypeError  :
#             number_of_SECTIONS = input( 'incorrect number of sections, introduce a numeric value: ' )
#
#
#     else:
#          value_of_the_midpoints = [ round( df.shape[0] / number_of_SECTIONS * (n+1) ) for n in range( number_of_SECTIONS ) ]
#     #               break
#          #     value_of_the_midpoints=[round(Mesh_points*n/(number_of_SECTIONS)+(Mesh_points/number_of_SECTIONS)) for n in range(number_of_SECTIONS)]
#
#
#
#     R2_train = {}
#     R2_train_Problematic = {}
#     R2_test = {}
#     R2_test_Problematic = {}
#     coefficients = {}
#     coefficients_Problematic = {}
#     Sections = []
#     cp = []
#     cp_problematic = []
#     #Full_shape = df[['x', 'y', 'z']]
#     #Full_shape.sort_values(by='x')
#     #Full_shape = Full_shape.drop_duplicates(inplace=False)
#     #Full_shape.reset_index()
#     #L = len(Full_shape)
#     keys=[]
#     print('Sections sorting Y')
#
#     pd.options.mode.chained_assignment = None  # default='warn'
#     for pos in np.arange( number_of_SECTIONS ):
#          keys.append('section '+str(pos))
#          if pos==0:
#               cuty=df[:value_of_the_midpoints[pos]]
#
#     #          Sections.append(cuty[['x','y','z']].drop_duplicates())#check how bad is it to drop duplicates, size 382.000-->16.200
#               Sections.append(cuty[['x','y','z']])
#
#               Sections[pos]['Mach'], Sections[pos]['Alpha'], Sections[pos]['q']= desired_params['Mach'][0], desired_params['Alpha'][0], desired_params['q'][0]
#
#     #          Full_shape['Mach'], Full_shape['Alpha'], Full_shape['q']= desired_Mach, desired_Alpha, desired_q
#     #          [coeff_cut , test_score_cut , train_score_cut,  TIME, cp_prediction] = modelito(df, poly_order, key, X_new)
#     #
#     #          if test_score_cut > 0:
#     #               R2_test [key] = test_score_cut
#     #               R2_train [key] = train_score_cut
#     #               coefficients [key] = coeff_cut
#     #               cp .append( cp_prediction)
#     #
#     #          else:
#     #               R2_test_Problematic [key] = test_score_cut
#     #               R2_train_Problematic [key] = train_score_cut
#     #               coefficients_Problematic [key] = coeff_cut
#     #               cp_problematic.append(cp_prediction)
#
#          else:
#               cuty=df[value_of_the_midpoints [pos-1] : value_of_the_midpoints [pos] ]
#
#               Sections.append(cuty[['x','y','z']].drop_duplicates())
#     #          Sections.append(cuty[['x','y','z']])
#
#               Sections[pos]['Mach'], Sections[pos]['Alpha'], Sections[pos]['q']= desired_params
#     #
#     #               X_new=Full_shape.iloc[ ((pos)*round(L/number_of_SECTIONS)): ((pos+1)*round(L/number_of_SECTIONS)) ]
#     #               X_new['Mach'], X_new['Alpha'], X_new['q']= desired_Mach, desired_Alpha, desired_q
#
#     [coeff_cut , test_score_cut , train_score_cut,  TIME, cp_prediction ] = modelito( df, poly_order, keys, Sections )
#
#     for pos,key in enumerate (keys):
#          if test_score_cut > 0:
#               R2_test [key] = test_score_cut
#               R2_train [key] = train_score_cut
#               coefficients [key] = coeff_cut
#               cp.append (cp_prediction[pos])
#
#          else:
#               R2_test_Problematic [key] = test_score_cut
#               R2_train_Problematic [key] = train_score_cut
#               coefficients_Problematic [key] = coeff_cut
#               cp_problematic.append(cp_prediction[pos])
#          #
#
#          #
#     #
#     #
#     #     R2_train_mean = np.array(list( R2_train_Problematic.values()) + list( R2_train.values() ) ).mean()
#     #     R2_test_mean = np.array(list( R2_test_Problematic.values()) + list( R2_test.values() ) ).mean()
#     R2_train_mean = sum( R2_train.values() ) / len( R2_train)
#     R2_test_mean = np.array( list(R2_test.values() )).mean()
#     print('mean R2_train value:    ', R2_train_mean )
#     print('mean R2_test value:     ', R2_test_mean )
#     end=time.time()
#     print('time modeling sections: ', end-start,'seconds \n')
#
#     try:
#          print(coefficients[key].head(10))
#     except KeyError:
#          print(coefficients_Problematic[key].head(10))
#     #
#     return  Sections, R2_test_mean, R2_test, R2_test_Problematic, R2_train_mean,  R2_train, R2_train_Problematic, coefficients, coefficients_Problematic, cp, cp_problematic
##
##
##
##
#[Sections, R2_test_mean, R2_test, R2_test_Problematic,  R2_train_mean,  R2_train, R2_train_Problematic, coefficients, coefficients_Problematic, cp, cp_problematic, Full_shape] = sections_y()
#FinalDF=pd.concat(Sections)
#FinalCP=np.concatenate(cp)

# ########################################### FULL MODEL  #############################################
#
#[coef_FM, test_score_FM, train_score_FM, TIME_FM, cp_predicted_FM ] = modelito( df.sort_values( by='x' , inplace=False), 3,'fullmodel',Full_shape)#add the geometry witout duplicated and with the flight parameterss

#scatter4d( Full_shape['x'], Full_shape['y'], Full_shape['z'], cp_predicted_FM)



#
#
#print('\nFull model \nR2_train: ', test_score_FM, '\nR2_test: ', train_score_FM)
#print('time for full model: ',TIME_FM, 'seg\n')

#coeff2= coeff1[(coeff1[['Coefficients Metamodel']] != 0).all(axis=1)]#drop zeros

# ##################################### Ploting Results ############################################

#cNorm = matplotlib.colors.Normalize(vmin=min(df['cp']), vmax=max(df['cp']))
##scalarMap = cmx.ScalarMappable(norm=cNorm, cmap=cm)
#fig = plt.figure()
#ax = Axes3D(fig)
##scalarMap.set_array(cp)
##fig.colorbar(scalarMap,label='Cp') ##DeltaP (daN/m2)
##    fig.suptitle(f, fontsize=8)
#ax.set_xlabel('x',fontsize=14, fontweight='bold')
#ax.set_ylabel('y',fontsize=14, fontweight='bold')
#ax.set_zlabel('z',fontsize=14, fontweight='bold')
#max_range = np.array([max(Full_shape['x'])-min(Full_shape['x']), max(Full_shape['y'])-min(Full_shape['y']), max(Full_shape['z'])-min(Full_shape['z'])]).max() / 2.0
#mid_x = (max( Full_shape['x'] ) + min( Full_shape['x'])) * 0.5
#mid_y = (max( Full_shape['y'] ) + min( Full_shape['y'])) * 0.5
#mid_z = (max( Full_shape['z'] ) + min( Full_shape['z'])) * 0.5
#ax.set_xlim(mid_x - max_range, mid_x + max_range)
#ax.set_ylim(mid_y - max_range, mid_y + max_range)
#ax.set_zlim(mid_z - max_range, mid_z + max_range)
#fig = plt.figure()


# ##################################### Ploting All cuts ############################################

#for n in range( len( Sections )):
##
###scatter4d( SectionsY[1]['x'], SectionsY[1]['y'], SectionsY[1]['z'], cp_problematic['section 1'])
#     section= 'section '+str(n)
#     scatter4d( Sections[n]['x'], Sections[n]['y'], Sections[n]['z'], cp[n],section)

# ##################################### Ploting Results ############################################



#scatter4d( FinalDF['x'], FinalDF['y'], FinalDF['z'], FinalCP,'full_cuts')












# ##############################################  KERAS   ###############################################

## Import the SGD optimizer
#import keras
#from keras.layers import Dense
#from keras.models import Sequential
#from keras.optimizers import SGD
#
## Create list of learning rates: lr_to_test
#LearningRate_to_test = [.000001, 0.01, 1]
#
## Loop over learning rates
#for lr in LearningRate_to_test:
#    print('\n\nTesting model with learning rate: %f\n'%lr )
#
#    # Build new model to test, unaffected by previous models
## Import necessary modules
#
#X   = df[['q','Alpha','Mach','z','y','x']]#'x','y','z'
#Y   = df[['cp']].values
## Specify the model
#n_cols = X.shape[1]
#model = Sequential()
#model.add(Dense(50, activation='relu', input_shape = (n_cols,)))
#model.add(Dense(32, activation='relu'))
#model.add(Dense(1))
#
## Compile the model
#model.compile(optimizer='adam', loss='mean_squared_error')
#
## Fit the model
#model.fit(X,Y)
##     #plt.show()
#


#


"""